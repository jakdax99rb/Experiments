from zipfile import ZipFile, ZIP_STORED, ZipInfo
import tempfile
import shutil
import os
import openpyxl
import re
import xlwings as xw

# This array stores what sheets are meant to be directly copied.
sheetsToBeCopied = ['Data', 'MedicationFull', 'DxListFull', 'VitalsHistory',
                    'ClientPhysicians', 'VisitDates', 'SnHistorical', 'RNVisitFormHistorical']
sheetPathsToBeCopied = ['xl/worksheets/sheet4.xml',
                        'xl/worksheets/sheet53.xml',
                        'xl/worksheets/sheet56.xml',
                        'xl/worksheets/sheet76.xml',
                        'xl/worksheets/sheet58.xml',
                        'xl/worksheets/sheet41.xml',
                        'xl/worksheets/sheet69.xml',
                        'xl/worksheets/sheet70.xml',
                        'xl/worksheets/_rels/sheet4.xml.rels',
                        'xl/worksheets/_rels/sheet53.xml.rels',
                        'xl/worksheets/_rels/sheet56.xml.rels',
                        'xl/worksheets/_rels/sheet76.xml.rels',
                        'xl/worksheets/_rels/sheet58.xml.rels',
                        'xl/worksheets/_rels/sheet41.xml.rels',
                        'xl/worksheets/_rels/sheet69.xml.rels',
                        'xl/worksheets/_rels/sheet70.xml.rels']

# Store master files in the same directory as this py file.
masterATL = "Nurse Admission File 9.6 MASTER.xlsm"
masterAugusta = "Nurse Admission File 9.6 Augusta MASTER.xlsm"


# call updater first to handle raw data transfer. Then call my zip updater after.
def updater(filePath, officeType):

    if officeType == 'augusta':

        masterPath = masterAugusta
    else:
        masterPath = masterATL

    sourceWorkBook = openpyxl.load_workbook(
        filePath, read_only=False, keep_vba=True, keep_links=False)
    destinationWorkbook = openpyxl.load_workbook(
        masterPath, read_only=False, keep_vba=True, keep_links=False)

    # Updates all sheets in the sheetsToBeCopied array

    for sheets in sheetsToBeCopied:

        sh = sourceWorkBook[sheets]
        mr = sh.max_row
        mc = sh.max_column

        for i in range(1, mr + 1):

            for j in range(1, mc + 1):

                c = sh.cell(row=i, column=j)

                # Checks if its a merged cell, if so skip.
                if(not re.search('MergedCell', str(destinationWorkbook[sheets].cell(row=i, column=j)))):

                    destinationWorkbook[sheets].cell(
                        row=i, column=j).value = c.value

    # Event History Management, this is when updating to 9.6

    sh = sourceWorkBook['EventHistory']
    mr = sh.max_row
    mc = sh.max_column

    for i in range(1, mr + 1):

        for j in range(1, mc + 1):

            c = sh.cell(row=i, column=j)

            # handles the changed columns
            if j == 1:

                destinationWorkbook['EventHistory'].cell(
                    row=i, column=3).value = c.value

            elif j == 3:

                destinationWorkbook['EventHistory'].cell(
                    row=i, column=21).value = c.value

            else:

                destinationWorkbook['EventHistory'].cell(
                    row=i, column=j).value = c.value

    # close and save openpyxl instances
    sourceWorkBook.close()
    destinationWorkbook.save(filePath.replace('9.3.xlsm', '9.6.xlsm'))
    destinationWorkbook.close()


'''
    xw.App(visible=False)

    # xlWings section for copying the dashboard over.
    print('made it to xlWings')

    # opens xlwings instances
    xlNewWorkBook = xw.Book(re.sub('9.3.xlsm', '9.6.xlsm', fileName))
    xlMasterWorkBook = xw.Book(newWorkBookPath)
    # deletes and re adds the dashboard

    sht = xlMasterWorkBook.sheets['Dashboard']
    xlNewWorkBook.sheets['Dashboard'].delete()
    sht.api.Copy(Before=xlNewWorkBook.sheets[1].api)

    # changes every cell in the dashboard to that of what it should be

    sh = xlNewWorkBook.sheets['Dashboard']
    sheetRange = sh.used_range
    mr = sheetRange.rows.count
    mc = sheetRange.columns.count

    for i in range(1, mr + 1):

        for j in range(1, mc + 1):

            c = sh.range((i, j))
            print(c.value)
            print(c)
            print('\n')

            newVersionWorkBook.sheets[sheets].range((i, j)).value = c.value

    xlNewWorkBook.save()
    xlMasterWorkBook.close()
    xlNewWorkBook.close()
    '''
# zipNurse(masterPath, filePath.replace('9.3.xlsm', '9.6.xlsm'))


def zipCopier(sourcePath, masterPath):

    targetFiles = []
    desitinationPath = sourcePath.replace(
        '9.3.xlsm', '9.6.xlsm')
    shutil.copyfile(masterPath, desitinationPath)

    with ZipFile(sourcePath, 'r') as sourceZip:

        for sheetPath in sheetPathsToBeCopied:

            targetFiles.append(sourceZip.read(sheetPath))

    with UpdateableZipFile(desitinationPath, 'a') as desitinationZip:

        for x in range(len(targetFiles)):

            desitinationZip.writestr(sheetPathsToBeCopied[x], targetFiles[x])


class UpdateableZipFile(ZipFile):
    """
    Add delete (via remove_file) and update (via writestr and write methods)
    To enable update features use UpdateableZipFile with the 'with statement',
    Upon  __exit__ (if updates were applied) a new zip file will override the exiting one with the updates
    """

    class DeleteMarker(object):
        pass

    def __init__(self, file, mode="r", compression=ZIP_STORED, allowZip64=False):
        # Init base
        super(UpdateableZipFile, self).__init__(file, mode=mode,
                                                compression=compression,
                                                allowZip64=allowZip64)
        # track file to override in zip
        self._replace = {}
        # Whether the with statement was called
        self._allow_updates = False

    def writestr(self, zinfo_or_arcname, bytes, compress_type=None):
        if isinstance(zinfo_or_arcname, ZipInfo):
            name = zinfo_or_arcname.filename
        else:
            name = zinfo_or_arcname
        # If the file exits, and needs to be overridden,
        # mark the entry, and create a temp-file for it
        # we allow this only if the with statement is used
        if self._allow_updates and name in self.namelist():
            temp_file = self._replace[name] = self._replace.get(name,
                                                                tempfile.TemporaryFile())
            temp_file.write(bytes)
        # Otherwise just act normally
        else:
            super(UpdateableZipFile, self).writestr(zinfo_or_arcname,
                                                    bytes, compress_type=compress_type)

    def write(self, filename, arcname=None, compress_type=None):
        arcname = arcname or filename
        # If the file exits, and needs to be overridden,
        # mark the entry, and create a temp-file for it
        # we allow this only if the with statement is used
        if self._allow_updates and arcname in self.namelist():
            temp_file = self._replace[arcname] = self._replace.get(arcname,
                                                                   tempfile.TemporaryFile())
            with open(filename, "rb") as source:
                shutil.copyfileobj(source, temp_file)
        # Otherwise just act normally
        else:
            super(UpdateableZipFile, self).write(filename,
                                                 arcname=arcname, compress_type=compress_type)

    def __enter__(self):
        # Allow updates
        self._allow_updates = True
        return self

    def __exit__(self, exc_type, exc_val, exc_tb):
        # call base to close zip file, organically
        try:
            super(UpdateableZipFile, self).__exit__(exc_type, exc_val, exc_tb)
            if len(self._replace) > 0:
                self._rebuild_zip()
        finally:
            # In case rebuild zip failed,
            # be sure to still release all the temp files
            self._close_all_temp_files()
            self._allow_updates = False

    def _close_all_temp_files(self):
        for temp_file in self._replace.values():
            if hasattr(temp_file, 'close'):
                temp_file.close()

    def remove_file(self, path):
        self._replace[path] = self.DeleteMarker()

    def _rebuild_zip(self):
        tempdir = tempfile.mkdtemp()
        try:
            temp_zip_path = os.path.join(tempdir, 'new.zip')
            with ZipFile(self.filename, 'r') as zip_read:
                # Create new zip with assigned properties
                with ZipFile(temp_zip_path, 'w', compression=self.compression,
                             allowZip64=self._allowZip64) as zip_write:
                    for item in zip_read.infolist():
                        # Check if the file should be replaced / or deleted
                        replacement = self._replace.get(item.filename, None)
                        # If marked for deletion, do not copy file to new zipfile
                        if isinstance(replacement, self.DeleteMarker):
                            del self._replace[item.filename]
                            continue
                        # If marked for replacement, copy temp_file, instead of old file
                        elif replacement is not None:
                            del self._replace[item.filename]
                            # Write replacement to archive,
                            # and then close it (deleting the temp file)
                            replacement.seek(0)
                            data = replacement.read()
                            replacement.close()
                        else:
                            data = zip_read.read(item.filename)
                        zip_write.writestr(item, data)
            # Override the archive with the updated one
            shutil.move(temp_zip_path, self.filename)
        finally:
            shutil.rmtree(tempdir)


def xlWingsTest(fileName, officeType):

    if officeType == 'augusta':

        newWorkBookPath = masterAugusta
    else:
        newWorkBookPath = masterATL

    xw.App(visible=True)

    sourceBook = xw.Book(fileName)
    destinationBook = xw.Book(newWorkBookPath)

    for sheets in sheetsToBeCopied:

        newWorkBook = sheetCopierXlwings(
            sourceBook, destinationBook, sheets)

    newWorkBook.save(fileName.replace('9.4.xlsm', '9.6.xlsm'))
    sourceBook.close()
    destinationBook.close()


def sheetCopierXlwings(sourceBook, destinationBook, sheets):

    if sheets != 'EventHistory':
        '''
        sht = workBook.sheets[sheets]
        sht.api.Copy(Before=newWorkBook.sheets[1].api)
        newWorkBook.sheets[sheets].api.Visible = False
        '''

        sh = sourceBook.sheets[sheets]
        sheetRange = sh.used_range
        mr = sheetRange.rows.count
        mc = sheetRange.columns.count

        for i in range(1, mr + 1):

            for j in range(1, mc + 1):

                c = sh.range((i, j))

                destinationBook.sheets[sheets].range((i, j)).value = c.value

    else:

        sh = sourceBook.sheets[sheets]
        sheetRange = sh.used_range
        mr = sheetRange.rows.count
        mc = sheetRange.columns.count

        for i in range(1, mr + 1):

            for j in range(1, mc + 1):

                c = sh.range((i, j))

                # handles the changed cell
                if j == 1:

                    destinationBook.sheets['EventHistory'].range(
                        (i, 3)).value = c.value

                elif j == 3:

                    destinationBook.sheets['EventHistory'].range(
                        (i, 21)).value = c.value

                else:

                    destinationBook.sheets[sheets].range(
                        (i, j)).value = c.value


xlWingsTest(
    r'D:\ProsperCareFiles\Editable\Givens, Jeanette Admission 11.14.2017 9.4.xlsm', r'Nurse Admission File 9.6 MASTER.xlsm')
